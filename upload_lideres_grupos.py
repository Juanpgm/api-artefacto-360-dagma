"""
Carga datos desde env/lideres_grupos_dagma.xlsx a Firestore.
Coleccion destino: lideres_grupos
"""
import argparse
import os
import sys
from typing import List

from openpyxl import load_workbook

# Agregar el directorio padre al path para importar modulos
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from app.firebase_config import db  # noqa: E402


def _normalize_headers(headers: List[object]) -> List[str]:
    normalized = []
    seen = {}
    for idx, header in enumerate(headers, start=1):
        key = "" if header is None else str(header)
        key = " ".join(key.split())
        if not key:
            key = f"col_{idx}"
        if key in seen:
            seen[key] += 1
            key = f"{key}_{seen[key]}"
        else:
            seen[key] = 1
        normalized.append(key)
    return normalized


def _row_to_dict(headers: List[str], row_values: List[object]) -> dict:
    data = {}
    for key, value in zip(headers, row_values):
        if value is None:
            continue
        if isinstance(value, str):
            value = value.strip()
            if not value:
                continue
        data[key] = value
    return data


def upload_excel_to_firestore(file_path: str, collection_name: str, sheet_name: str | None, limit: int | None) -> None:
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"No se encontro el archivo: {file_path}")

    wb = load_workbook(file_path, data_only=True)
    sheet = wb[sheet_name] if sheet_name else wb.active

    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers_row = next(rows_iter)
    except StopIteration:
        raise ValueError("El archivo no tiene filas")

    headers = _normalize_headers(list(headers_row))

    total_uploaded = 0
    total_skipped = 0
    batch = db.batch()
    batch_count = 0

    for row_index, row_values in enumerate(rows_iter, start=2):
        if limit is not None and total_uploaded >= limit:
            break
        data = _row_to_dict(headers, list(row_values))
        if not data:
            total_skipped += 1
            continue
        doc_ref = db.collection(collection_name).document()
        batch.set(doc_ref, data)
        batch_count += 1
        total_uploaded += 1
        if batch_count >= 500:
            batch.commit()
            batch = db.batch()
            batch_count = 0

    if batch_count > 0:
        batch.commit()

    print("Carga finalizada")
    print(f"Coleccion: {collection_name}")
    print(f"Archivo: {file_path}")
    if sheet_name:
        print(f"Hoja: {sheet_name}")
    print(f"Registros cargados: {total_uploaded}")
    print(f"Filas omitidas (vacias): {total_skipped}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Carga lideres_grupos_dagma.xlsx a Firestore")
    parser.add_argument(
        "--file",
        default=os.path.join("env", "lideres_grupos_dagma.xlsx"),
        help="Ruta al archivo XLSX",
    )
    parser.add_argument(
        "--collection",
        default="lideres_grupos",
        help="Nombre de la coleccion en Firestore",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Nombre de la hoja a usar (por defecto la hoja activa)",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Limite de filas a cargar (opcional)",
    )

    args = parser.parse_args()
    upload_excel_to_firestore(args.file, args.collection, args.sheet, args.limit)


if __name__ == "__main__":
    main()
